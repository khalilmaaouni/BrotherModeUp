#!/usr/bin/env python3
"""The OPTIONAL exporter for the documentation folder: PDF, Word, Excel.

READ THIS FIRST
  This module is optional in the strongest sense available: nothing imports it.
  tools/bm_docs.py, the mandatory generator, does not import it, does not
  mention it in a lazy import, and does not degrade if this file is deleted. The
  only way it runs is a human running it. That is deliberate, and it is what
  keeps section 5.6's promise structural rather than a comment: an optional
  dependency reachable from the mandatory path is a dependency.

  Phase B of docs/superpowers/specs/2026-07-30-documentation-and-gate-packs
  -design.md, section 5.6.

WHAT "TOOLING ALREADY ON THE MACHINE" MEANS HERE
  Importable Python modules that are ALREADY installed. Not a subprocess: every
  shipping tool in tools/ is forbidden from running one (I3), so this cannot
  shell out to pandoc, libreoffice or wkhtmltopdf even where they exist, and
  pretending otherwise would be the kind of claim this project's own suite
  fails a document for. Nothing is installed, nothing is downloaded, and no
  entry is added to any dependency list (I2).

  So the honest set is: a PDF writer, a Word writer or an Excel writer that the
  founder already has in this interpreter. When one is absent this says which
  format could not be produced AND WHY, in one line a human can act on, and
  exits 0 for a report and 1 only when the caller asked for a format that could
  not be produced. It never raises into anything, because a missing optional
  format is not an error in the documentation.

WHAT IT WILL NEVER DO
  Silently produce a worse file under a name that claims a format. A .pdf that
  is really text, or an .xlsx that is really a csv, is worse than an absence: a
  reader opens it, it works badly, and nobody knows why. Markdown plus mermaid
  is the format this project always produces (founder decision 3), and it is
  already on disk.

Python 3.9, standard library only.

No em or en dashes anywhere in this file or its output.

Usage:
  python3 tools/bm_docs_export.py report              # what could be produced
  python3 tools/bm_docs_export.py export              # produce what is possible
  python3 tools/bm_docs_export.py export --format pdf # and fail if it cannot
"""

import importlib.util
import io
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))

# The generated folder this exports FROM. Duplicated as a literal rather than
# imported from bm_docs, because importing the generator to read one constant
# would create the edge this module exists to not have.
DOC_ROOT = "Documentation"

# format -> (candidate module names, what it would be used for, whether this
# project ships a WRITER for it). Ordered: the first importable candidate wins.
#
# TWO SEPARATE QUESTIONS, KEPT SEPARATE ON PURPOSE. Whether a backend is
# importable on this machine is one fact. Whether BrotherMode ships a tested
# writer against that backend is a different one, and collapsing them is how an
# exporter comes to claim a format it has never actually produced. Only xlsx has
# a writer today, and it has one because it could be exercised end to end here;
# pdf and docx say so rather than shipping a code path nobody has run.
BACKENDS = (
    ("pdf", ("reportlab", "fpdf", "weasyprint"),
     "render the markdown pages as a paginated document", False),
    ("docx", ("docx",),
     "write the pages as a Word document", False),
    ("xlsx", ("openpyxl", "xlsxwriter"),
     "write the work breakdown and schedule tables as a workbook", True),
)

FORMATS = tuple(name for name, _c, _w, _s in BACKENDS)

# Where the machine-readable facts live. Read rather than recomputed: the
# exporter must not become a second projection of the store, because then two
# things could disagree about the same number.
FACTS_REL = (DOC_ROOT, "90-generated", "facts.json")

# Where a produced workbook lands.
EXPORT_DIR = (DOC_ROOT, "90-generated")


def _out(msg=""):
    sys.stdout.write("%s\n" % msg)


def _err(msg):
    sys.stderr.write("%s\n" % msg)


def _importable(name):
    """Is this module importable WITHOUT importing it.

    find_spec rather than a try/import: importing a heavy optional library to
    discover whether it exists costs the import even when the answer is not
    needed, and any side effect it has on import belongs to the founder's
    decision to use it, not to this probe."""
    try:
        return importlib.util.find_spec(name) is not None
    except (ImportError, ValueError, AttributeError):
        # A namespace package or a broken installation can raise here. Treated
        # as absent, which is the truthful answer: it cannot be used.
        return False


def probe():
    """One honest row per format: can it be produced, by what, and if not why.

    This is the whole of section 5.6's contract. It reports; it does not
    apologize and it does not guess."""
    rows = []
    for fmt, candidates, what, shipped in BACKENDS:
        found = ""
        for name in candidates:
            if _importable(name):
                found = name
                break
        if not found:
            rows.append({
                "format": fmt, "available": False, "backend": "",
                "writer_shipped": shipped,
                "reason": "no writer for %s is installed here. Tried %s, none "
                          "importable. It would have been used to %s. "
                          "BrotherMode adds no dependency to get it: install "
                          "one yourself if you want this format, or use the "
                          "markdown that is already on disk."
                          % (fmt, ", ".join(candidates), what)})
        elif not shipped:
            rows.append({
                "format": fmt, "available": False, "backend": found,
                "writer_shipped": False,
                "reason": "%s IS importable here, so this format is possible on "
                          "this machine, and BrotherMode still ships no writer "
                          "for it: the conversion would be a backend specific "
                          "code path that no test in this project can exercise "
                          "on a machine without %s, and an untested writer "
                          "producing a document somebody hands to a regulator "
                          "is worse than an honest gap. Point %s at the markdown "
                          "in %s yourself."
                          % (found, found, found, DOC_ROOT)})
        else:
            rows.append({
                "format": fmt, "available": True, "backend": found,
                "writer_shipped": True,
                "reason": "%s is importable in this interpreter and is used to "
                          "%s" % (found, what)})
    return rows


def sources(root):
    """The generated markdown pages, repo-relative and sorted. Empty when the
    folder has not been generated yet, which is a different problem and gets a
    different message."""
    base = os.path.join(root, DOC_ROOT)
    out = []
    for dirpath, dirnames, filenames in os.walk(base):
        dirnames[:] = sorted(dirnames)
        for name in sorted(filenames):
            if name.endswith(".md"):
                full = os.path.join(dirpath, name)
                out.append(os.path.relpath(full, root).replace("\\", "/"))
    return sorted(out)


def _read(root, rel):
    with io.open(os.path.join(root, rel), encoding="utf-8",
                 errors="replace") as fh:
        return fh.read()


def _facts(root):
    """The generated facts, or None. Never raises: a folder that has not been
    generated, or a facts file somebody truncated, is a message rather than a
    traceback."""
    path = os.path.join(root, *FACTS_REL)
    if not os.path.isfile(path):
        return None
    try:
        with io.open(path, encoding="utf-8") as fh:
            data = json.load(fh)
    except (ValueError, OSError):
        return None
    return data if isinstance(data, dict) else None


def _sheets(facts):
    """The tables worth a spreadsheet, as (sheet name, header row, data rows).

    The work breakdown and the schedule, because those are the two things people
    ask for in a spreadsheet: a list of work with its state, and the chain that
    decides the finish date. Everything else in the folder is prose, and prose in
    a cell is prose nobody reads."""
    items = facts.get("work_items") or []
    schedule = facts.get("schedule") or {}
    nodes = schedule.get("nodes") or []
    edges = schedule.get("edges") or []
    path = schedule.get("critical_path") or []
    wbs_rows = []
    for i, item in enumerate(items, 1):
        wbs_rows.append([
            i, item.get("short", ""), item.get("name", ""),
            item.get("state", ""), item.get("lifetime", ""),
            item.get("owner", ""), item.get("objective", ""),
            len(item.get("files") or []), len(item.get("decisions") or []),
            item.get("created_at", ""), item.get("updated_at", "")])
    sched_rows = []
    for node in nodes:
        sched_rows.append([
            node.get("id", ""), node.get("name", ""), node.get("state", ""),
            node.get("start", ""), node.get("days", 0),
            "yes" if node.get("id") in path else "no"])
    dep_rows = []
    for edge in edges:
        pair = list(edge) + ["", "", ""]
        dep_rows.append([pair[0], pair[1], pair[2]])
    return (
        ("WBS", ["#", "record", "name", "state", "lifetime", "owner",
                 "objective", "files claimed", "decisions", "opened",
                 "last touched"], wbs_rows),
        ("Schedule", ["record", "name", "state", "start", "recorded days",
                      "on critical path"], sched_rows),
        ("Dependencies", ["must precede", "waits for it", "because both claim"],
         dep_rows),
    )


def _write_xlsx(path, sheets, backend):
    """Produce a workbook with whichever backend was found.

    Two backends because both are common and neither is a dependency. Each is
    imported HERE, inside the one function that uses it, so importing this module
    imports neither."""
    if backend == "openpyxl":
        import openpyxl
        book = openpyxl.Workbook()
        book.remove(book.active)
        for name, header, rows in sheets:
            sheet = book.create_sheet(title=name)
            sheet.append(list(header))
            for row in rows:
                sheet.append(list(row))
        book.save(path)
        return
    if backend == "xlsxwriter":
        import xlsxwriter
        book = xlsxwriter.Workbook(path)
        for name, header, rows in sheets:
            sheet = book.add_worksheet(name)
            for col, title in enumerate(header):
                sheet.write(0, col, title)
            for r, row in enumerate(rows, 1):
                for col, value in enumerate(row):
                    sheet.write(r, col, value)
        book.close()
        return
    raise ValueError("no xlsx backend named %r" % backend)


def export(root, wanted=None):
    """Produce what can be produced, and state exactly what could not.

    Returns a report. Raises nothing: a missing optional format is a line in the
    report, and a caller who demanded a specific format learns it from the
    report's `refused` list rather than from an exception. A backend that is
    present and then FAILS is also a line in the report, carrying the backend's
    own error text, because "it did not work" and "you do not have it" are
    different problems with different fixes."""
    pages = sources(root)
    facts = _facts(root)
    rows = probe()
    report = {"pages": len(pages), "produced": [], "refused": [],
              "backends": rows, "notes": []}
    if not pages:
        report["notes"].append(
            "%s holds no markdown page yet, so there is nothing to export. Run "
            "the generator first: bm_docs.py generate." % DOC_ROOT)
    for row in rows:
        fmt = row["format"]
        if wanted and fmt not in wanted:
            continue
        if not row["available"]:
            report["refused"].append({"format": fmt, "why": row["reason"]})
            continue
        if facts is None:
            report["refused"].append({
                "format": fmt,
                "why": "a %s writer is available (%s) but %s carries no "
                       "readable facts.json, so there is nothing to put in it. "
                       "Run bm_docs.py generate first."
                       % (fmt, row["backend"], DOC_ROOT)})
            continue
        directory = os.path.join(root, *EXPORT_DIR)
        target = os.path.join(directory, "project-tables.%s" % fmt)
        try:
            if not os.path.isdir(directory):
                os.makedirs(directory)
            _write_xlsx(target, _sheets(facts), row["backend"])
        except Exception as exc:
            # Deliberately broad. Every one of these backends is third party
            # code this project does not test, and whatever it raises must
            # become a stated limitation rather than a traceback out of an
            # OPTIONAL step. The backend's own words are carried through.
            report["refused"].append({
                "format": fmt,
                "why": "%s is installed but failed while writing %s: %s: %s"
                       % (row["backend"], target, type(exc).__name__, exc)})
            continue
        report["produced"].append({
            "format": fmt, "backend": row["backend"],
            "path": os.path.relpath(target, root).replace("\\", "/")})
    return report


def _root():
    """The project root, resolved the way every other tool resolves it. Imported
    lazily and by path so that this module has no import-time edge to the
    toolchain at all: `report` works in a directory with no store."""
    spec = importlib.util.spec_from_file_location(
        "bm_store_for_export", os.path.join(HERE, "bm_store.py"))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    root, _source = mod.require_root()
    return root


def cmd_report(argv):
    as_json = "--json" in argv
    rows = probe()
    if as_json:
        _out(json.dumps({"backends": rows}, indent=2, sort_keys=True))
        return 0
    _out("optional export formats on this machine:")
    for row in rows:
        _out("  %-5s %s" % (row["format"],
                            "available via %s" % row["backend"]
                            if row["available"] else "NOT available"))
        _out("        %s" % row["reason"])
    _out("markdown plus mermaid is always produced and is already in %s."
         % DOC_ROOT)
    return 0


def cmd_export(argv):
    as_json = "--json" in argv
    wanted = None
    if "--format" in argv:
        i = argv.index("--format")
        if i + 1 >= len(argv):
            _err("bm_docs_export: --format needs a value (%s)"
                 % ", ".join(FORMATS))
            return 2
        wanted = [argv[i + 1]]
        if wanted[0] not in FORMATS:
            _err("bm_docs_export: unknown format %r (known: %s)"
                 % (wanted[0], ", ".join(FORMATS)))
            return 2
    report = export(_root(), wanted=wanted)
    if as_json:
        _out(json.dumps(report, indent=2, sort_keys=True))
    else:
        for note in report["notes"]:
            _out(note)
        for row in report["produced"]:
            _out("produced %s at %s via %s"
                 % (row["format"], row["path"], row["backend"]))
        for row in report["refused"]:
            _out("could NOT produce %s: %s" % (row["format"], row["why"]))
        _out("%d markdown page(s) are on disk in %s and need no exporter."
             % (report["pages"], DOC_ROOT))
    # Exit 1 only when the caller NAMED a format and did not get it. A bare
    # export is a report and a report that ran is a success.
    if wanted and report["refused"]:
        return 1
    return 0


COMMANDS = {"report": cmd_report, "export": cmd_export}


def main(argv):
    if not argv or argv[0] in ("-h", "--help", "help"):
        _out(__doc__.strip())
        _out("")
        _out("commands: %s" % ", ".join(sorted(COMMANDS)))
        return 0
    cmd = argv[0]
    if cmd not in COMMANDS:
        _err("bm_docs_export: unknown command %r (known: %s)"
             % (cmd, ", ".join(sorted(COMMANDS))))
        return 2
    return COMMANDS[cmd](argv[1:])


def cli():
    """Console-script entry point for a packaged install. Must be callable with
    no arguments, so it wraps main(argv) rather than being it."""
    sys.exit(main(sys.argv[1:]))


if __name__ == "__main__":
    cli()
